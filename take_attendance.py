import cv2
import json
import threading
import numpy as np
from pyzbar.pyzbar import decode
from playsound import playsound
import openpyxl
import datetime as dt
import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
from concurrent.futures import ThreadPoolExecutor

record = cv2.VideoCapture(0)
record.set(3,640)
record.set(4,480)

print("reading students ids ......")
xlsx_data = openpyxl.load_workbook("data.xlsx")
sheet = xlsx_data.active
max_rows = sheet.max_row

# Load student names from CSV (Placeholder)
id_to_name = {}  # Load this mapping from a CSV.
id_serials = {}
id_to_sections = {}

for i in range(1,max_rows+1):
    temp=sheet.cell(i,1)
    temp2=sheet.cell(i,2)
    id_to_name[temp.value] = temp2.value
    temp3 = sheet.cell(i,3)
    id_to_sections[temp.value] = temp3.value
    id_serials[temp2.value] = i

print("reading students free slots ......")
# Load student free slots (Placeholder)
f = open("FreeSlots.json")
free_slots = json.load(f)  # Load student free slots from a CSV.
f.close()

recorded_data = {}  # A dictionary to store In-time, Out-time, and LAN cable ID.
lan_data = {}  # A dictionary to store which student has which LAN cable.

last_student = None


# print("connecting to google sheets ......")
# #online Google Sheets
# scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
# creds = ServiceAccountCredentials.from_json_keyfile_name('attendance.json', scope)
# client = gspread.authorize(creds)
# wb = client.open('MERL v2.0 Attendance Sheet')
# sheet = wb.get_worksheet(-1)

# #column = sheet.cell(100,2).value
# max_row = len(sheet.get_all_values())
# max_col = len(sheet.get_all_values()[0])

# #print(column,type(column))
# #max_col=int(column)
# #print(max_col,type(max_col))

# to_day = str(dt.date.today())
# #print(to_day,type(to_day))

# sheet.update_cell(1,max_col+1,to_day)

def merge_cells(worksheet, start_row, start_col, end_row, end_col):
    """Merge cells in a given worksheet."""
    grid_range = {
        "startRowIndex": start_row - 1,
        "endRowIndex": end_row,
        "startColumnIndex": start_col - 1,
        "endColumnIndex": end_col
    }
    request = {
        "mergeCells": {
            "range": grid_range,
            "mergeType": "MERGE_ALL"
        }
    }
    body = {
        "requests": [request]
    }
    worksheet.spreadsheet.batch_update(body)
def preprocess_sheet():
    print("connecting to google sheets ......")
    #online Google Sheets
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('attendance.json', scope)
    client = gspread.authorize(creds)
    wb = client.open('MERL v2.0 Attendance Sheet')
    sheet = wb.get_worksheet(-1)

    #column = sheet.cell(100,2).value
    max_row = len(sheet.get_all_values())
    max_col = len(sheet.get_all_values()[0])

    #print(column,type(column))
    #max_col=int(column)
    #print(max_col,type(max_col))

    to_day = str(dt.date.today())
    date_columns = {sheet.cell(1, col).value: col for col in range(2, max_col + 1,4)}
    # print(f"date columns {date_columns}")
    print("preprocessing sheet ......")

    # Check if we need a new month's tab
    current_month = dt.date.today().strftime('%B %Y')
    if current_month not in [worksheet.title for worksheet in wb.worksheets()]:
        print("adding new month tab ......")
        new_worksheet=wb.add_worksheet(title=current_month, rows="100", cols="20")
        sheet = wb.get_worksheet(-1)  # Change the sheet to the newly created one
        # You'd need to set up the initial columns, headers, etc for this new sheet as well.
        for row in range(2, max_row + 1):
            student_name = sheet.cell(row, 1).value
            new_worksheet.update_cell(row, 1, student_name)

    if to_day not in date_columns:
        print("adding new day ......")
        new_col = max_col + 1
        sheet.update_cell(1, new_col, to_day)
        merge_cells(sheet, 1, new_col, 1, new_col + 4)
        sheet.update_cell(2, new_col + 0, "In-Time")
        sheet.update_cell(2, new_col + 1, "LAN ID")
        sheet.update_cell(2, new_col + 2, "Out-Time")
        sheet.update_cell(2, new_col + 3, "LAN Turn-In")

    else:
        new_col = date_columns[to_day]

threading.Thread(target=preprocess_sheet)

current_student = None


def update_cells_in_thread(i, col, value):
    print("connecting to google sheets ......")
    #online Google Sheets
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('attendance.json', scope)
    client = gspread.authorize(creds)
    wb = client.open('MERL v2.0 Attendance Sheet')
    sheet = wb.get_worksheet(-1)

    #column = sheet.cell(100,2).value
    max_row = len(sheet.get_all_values())
    max_col = len(sheet.get_all_values()[0])

    #print(column,type(column))
    #max_col=int(column)
    #print(max_col,type(max_col))
    to_day = str(dt.date.today())
    date_columns = {sheet.cell(1, col).value: col for col in range(2, max_col + 1,4)}


    if to_day not in date_columns:
        print("adding new day ......")
        new_col = max_col + 1
        sheet.update_cell(1, new_col, to_day)
        merge_cells(sheet, 1, new_col, 1, new_col + 4)
        sheet.update_cell(2, new_col + 0, "In-Time")
        sheet.update_cell(2, new_col + 1, "LAN ID")
        sheet.update_cell(2, new_col + 2, "Out-Time")
        sheet.update_cell(2, new_col + 3, "LAN Turn-In")

    else:
        new_col = date_columns[to_day]
    sheet.update_cell(i, new_col + col, value)

def check_value_in_sheet(i,j):
    print("connecting to google sheets ......")
    #online Google Sheets
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('attendance.json', scope)
    client = gspread.authorize(creds)
    wb = client.open('MERL v2.0 Attendance Sheet')
    sheet = wb.get_worksheet(-1)

    #column = sheet.cell(100,2).value
    max_row = len(sheet.get_all_values())
    max_col = len(sheet.get_all_values()[0])

    to_day = str(dt.date.today())
    date_columns = {sheet.cell(1, col).value: col for col in range(2, max_col + 1,4)}


    if to_day not in date_columns:
        print("adding new day ......")
        new_col = max_col + 1
        sheet.update_cell(1, new_col, to_day)
        merge_cells(sheet, 1, new_col, 1, new_col + 4)
        sheet.update_cell(2, new_col + 0, "In-Time")
        sheet.update_cell(2, new_col + 1, "LAN ID")
        sheet.update_cell(2, new_col + 2, "Out-Time")
        sheet.update_cell(2, new_col + 3, "LAN Turn-In")

    else:
        new_col = date_columns[to_day]

    heheh = sheet.cell(i,new_col + j).value
    print(heheh)
    return heheh


print("starting scan ......")

while(1):
    s, img = record.read()
    for qr in decode(img):
        data=qr.data.decode('utf-8')
        #print(data)

        #box around
        pts = np.array([qr.polygon],np.int32)
        pts = pts.reshape((-1,1,2))
        
        #display on image
        pts2 = qr.rect

        print(f"scanned {data} ......")
        if data.startswith("MERL"):
            nameOfStudent = id_to_name[data]
            # for i in range(2,max_row+1):
                #  x=sheet.cell(i,1)
            sr = id_serials[nameOfStudent] + 2 # row
                #  if(x.value == nameOfStudent):
            current_student = nameOfStudent
            # Using ThreadPoolExecutor
            with ThreadPoolExecutor() as executor:
                future = executor.submit(check_value_in_sheet, sr,0)
                sheet_value = future.result()
            print(sheet_value)
            if sheet_value == None:
                print(f"marking in-time of {nameOfStudent} ......")
                in_time = dt.datetime.now().strftime('%I:%M %p')
                print(f"in-time: {in_time}")
                threading.Thread(target=update_cells_in_thread, args=(sr, 0, in_time)).start()
                threading.Thread(target=update_cells_in_thread, args=(sr, 1, "no")).start()
                threading.Thread(target=update_cells_in_thread, args=(sr, 3, "no")).start()
                break
            else:   
                print(f"marking out-time of {nameOfStudent} ......")
                out_time = dt.datetime.now().strftime('%I:%M %p')
                print(f"out-time: {out_time}")
                threading.Thread(target=update_cells_in_thread, args=(sr, 2, out_time)).start()
                break

        elif data.startswith("LAN"):
            
            sr = id_serials[current_student] + 2 # row
            with ThreadPoolExecutor() as executor:
                future = executor.submit(check_value_in_sheet, sr,1)
                sheet_value = future.result()
            if sheet_value == "no":
                print(f"assigning lan {data} to {current_student} ......")
                # lan_taken = "yes"
                threading.Thread(target=update_cells_in_thread, args=(sr,  1, data)).start()
            else:
                print(f"lan {data} taken back from {current_student} ......")
                threading.Thread(target=update_cells_in_thread, args=(sr, 3, "yes")).start()

        
        
    cv2.imshow('Result',img)
    key = cv2.waitKey(1)

    #to release camera frame
    if(key==27): #esc key value is 27
        cv2.destroyWindow("Result")
        record.release()
        break
    
