import openpyxl
import qrcode
from mail_attachment import send_mail

#data_collection_to generate content in qr
data = openpyxl.load_workbook("data.xlsx")
sheet = data.active
#print(sheet)
names=[]
mails=[]

for i in range(1,54):
    temp=sheet.cell(i,1)
    names.append(temp.value)
    # temp2=sheet.cell(i,2)
    # mails.append(temp2.value)
        

print(names)
print(mails)

#qr code generation

for i in names:
    img = qrcode.make(i)
    img.save('QR/{}.png'.format(i))

#sending mail

# for i in range(len(names)):
#     send_mail(mails[i],names[i])

#This is a one time process
